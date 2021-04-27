from openpyxl import Workbook
from openpyxl.styles import Font, Alignment


class Reporter:
    def __init__(self, scanner, vulnerabilities, name, current_scan=None, previous_scan=None, site_id=None):
        self.vulnerabilities = vulnerabilities
        self.scanner = scanner
        self.name = name
        self.current_scan = current_scan
        self.previous_scan = previous_scan
        self.site_id = site_id

    def writeToXlsx(self, SD=None, dont_touch=None) -> None:
        scanner = self.scanner
        vulnerabilities = self.vulnerabilities
        if not self.current_scan or not self.previous_scan or not self.site_id:
            last_scan_site = scanner.last_scan_site_id()
            site_id = last_scan_site[1]
            last_scan = scanner.last_scan_by_site_id(site_id)
            assets_scan = scanner.assets_by_scan(last_scan[0])
            nodes_scan = scanner.nodes_by_scan(last_scan[1])
        else:
            assets_scan = scanner.assets_by_scan(self.current_scan)
            nodes_scan = scanner.nodes_by_scan(self.previous_scan)
            site_id = self.site_id
        vuln_ids = vulnerabilities.vuln_by_site(site_id)
        xlsx = Workbook()
        xlsx.create_sheet(title='Vulnerabilities', index=0)
        sheet = xlsx['Vulnerabilities']
        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 100
        sheet.column_dimensions['C'].width = 50
        sheet.column_dimensions['D'].width = 42
        sheet['A1'], sheet['B1'], sheet['C1'], sheet['D1'] = 'Vulnerability', 'Descriptions', 'Solutions', 'Asset'
        sheet.auto_filter.ref = 'A1:D1'
        sheet['A1'].font = sheet['B1'].font = sheet['C1'].font = sheet['D1'].font = Font(bold=True)
        sheet['A1'].alignment = sheet['B1'].alignment = sheet['C1'].alignment = Alignment(horizontal="center")
        num, num_list = 0, [2]
        for i in vuln_ids:
            sheet[f'A{num + 2}'] = vuln_ids[i][0]
            sheet[f'B{num + 2}'] = vulnerabilities.vuln_description(i)
            sheet[f'A{num + 2}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sheet[f'B{num + 2}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            # include exploit
            if vuln_ids[i][1]:
                sheet[f'A{num + 2}'].font = Font(color='4169E1', bold=True)
            # integration with it-assets (exclude dont_touch assets)
            if SD and dont_touch:
                keys_asset_vuln = vulnerabilities.asset_vulners(i, assets_scan)
                assets_without_critical_system = dict()
                for asset in keys_asset_vuln:
                    parent = SD.parent_ius(asset)
                    if not set(parent.split(' ')).intersection(dont_touch):
                        assets_without_critical_system[asset] = keys_asset_vuln[asset]
            else:
                assets_without_critical_system = vulnerabilities.asset_vulners(i, assets_scan)
            if assets_without_critical_system == {}:
                sheet[f'A{num + 2}'] = None
                sheet[f'B{num + 2}'] = None
                continue
            else:
                for j in assets_without_critical_system:
                    sheet[f'D{num + 2}'] = f'{j} ({assets_without_critical_system[j]})'
                    try:
                        if i in vulnerabilities.vuln_by_node(nodes_scan[j]):
                            sheet[f'D{num + 2}'].font = Font(color='FF0000', bold=True)
                    except KeyError:
                        continue
                    finally:
                        num += 1
            num_list.append(num + 2)
        for i in range(len(num_list) - 1):
            sheet.merge_cells(f'A{num_list[i]}:A{num_list[i + 1] - 1}')
            sheet.merge_cells(f'B{num_list[i]}:B{num_list[i + 1] - 1}')
            sheet.merge_cells(f'C{num_list[i]}:C{num_list[i + 1] - 1}')
        xlsx.save(f'{self.name}.xlsx')

    def create_report(self, report_type, SD=None, dont_touch=None) -> None:
        if report_type == 'xlsx':
            self.writeToXlsx(SD, dont_touch)
        else:
            print('unknown type')
